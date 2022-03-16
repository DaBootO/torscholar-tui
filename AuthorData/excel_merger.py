import openpyxl
import os
import string

uppercase_letters = string.ascii_uppercase

originals = ["PROCESS", "MATERIALS", "JOINING"]
new_ones = "_JAN_UPDATE"

base_directory = "/home/dabooto/PycharmProjects/AuthorData/Rohdaten/"

for original in originals:
    files_original = os.listdir(base_directory+original)
    files_new = os.listdir(base_directory+original+new_ones)
    for file in files_original:
        if file in files_new:
            print(file, "IS IN BOTH!")
            wb1 = openpyxl.load_workbook(base_directory+original+"/"+file)
            wb2 = openpyxl.load_workbook(base_directory+original+new_ones+"/"+file)

            ws1 = wb1.active
            ws2 = wb2.active

            index1 = 1
            max_row1 = ws1.max_row
            while ws1['C'+str(index1)].value != 'title' and index1 <= max_row1:
                index1 += 1
            index1 += 1
            index2 = 1
            max_row2 = ws2.max_row
            while ws2['C'+str(index2)].value != 'title' and index2 <= max_row2:
                index2 += 1
            index2 += 1

            print("index:", index1, "\t max_row:", max_row1)

            if index2 < max_row2:
                iterator = 0
                for row in range(index2, max_row2):
                    for col in range(12):
                        ws1.cell(row=(max_row1+1+iterator), column=col+1).value = ws2.cell(row=(index2+iterator), column=col+1).value
                    iterator += 1
                    # print(iterator)
            wb1.save(base_directory + original + "/" + file)
            wb2.save(base_directory + original + new_ones + "/" + file)

