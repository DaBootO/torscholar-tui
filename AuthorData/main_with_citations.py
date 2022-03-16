import os
import openpyxl
import re
import pickle


def save_obj(obj, name):
    with open('obj/'+ name + '.pkl', 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)


def load_obj(name):
    with open('obj/' + name + '.pkl', 'rb') as f:
        return pickle.load(f)


def dim2list(ws):
    dims = ws.dimensions.split(':')
    for i in range(len(dims)):
        dims[i] = int(re.findall('\d+', dims[i])[0])
    return dims

base_directory = "/home/dabooto/PycharmProjects/AuthorData/Rohdaten"

sub_dirs = os.listdir(base_directory)
print(sub_dirs)

file_number = 0
DataOut = {}
for sd in sub_dirs:
    if "." not in sd:
        DataOut[sd] = {}
        sub_dir_path = base_directory + '/' + sd
        files = os.listdir(sub_dir_path)
        for f in files:
            authors = {}

            if f[-4:] == 'xlsx':
                wb = openpyxl.load_workbook(sub_dir_path+'/'+f)
                ws = wb.active
                dims = dim2list(ws)

                row = 1
                while ws['D'+str(row)].value != 'author':
                    row += 1

                for row_a in range(row+1, dims[1]+1):
                    line = ws['D'+str(row_a)].value
                    is_part_of = ws['A'+str(row_a)].value
                    comment = ws['B'+str(row_a)].value
                    citations = ws['G'+str(row_a)].value
                    if type(citations) is str and citations != "None":
                        try:
                            citations = int(citations)
                        except Exception:
                            print(f, row_a)
                            print("something went wrong in the aforementioned sheet")
                            print(type(citations))
                    else:
                        citations = 0

                    if line is not None:
                        if is_part_of == 1 and comment != "doppelt":
                            if '-' in line:
                                auth = line[:line.index('-')]
                            else:
                                auth = line
                        else:
                            continue
                    else:
                        continue

                    #
                    # if line != None and '-' in line:
                    #     auth = line[:line.index('-')]
                    # else:
                    #     if line != None and ',' in line:
                    #         auth = line
                    #     else:
                    #         continue

                    auth = auth.strip().upper()
                    auth = re.sub(r'[^A-Z\s,]+', '', auth)
                    auth = auth.split(',')
                    for authi in auth:
                        authi = authi.strip()
                        if authi != "":
                            if authi in authors:
                                authors[authi][0] += 1
                                authors[authi][1] += citations
                            else:
                                authors[authi] = [1, citations]
                filename = f[:-5]
                DataOut[sd][filename] = authors
                wb.close()

print(DataOut)

save_obj(DataOut, "AuthorOutput")