import pickle
from openpyxl import Workbook
import string


def save_obj(obj, name):
    with open('obj/'+ name + '.pkl', 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)


def load_obj(name):
    with open('obj/' + name + '.pkl', 'rb') as f:
        return pickle.load(f)


uppercase_letters = list(string.ascii_uppercase)
Dict = load_obj("AuthorOutput")
output = {}


for theme in Dict.keys():
    author_dict = {}
    for query in Dict[theme].keys():
        for author in Dict[theme][query].keys():
            if author in list(author_dict.keys()):
                author_dict[author] += Dict[theme][query][author]
            else:
                author_dict[author] = Dict[theme][query][author]
    output[theme] = dict(sorted(author_dict.items(), key=lambda item: item[1], reverse=True))

wb = Workbook()
ws = wb.active

col_num = 0
for theme in output.keys():
    row_num = 1
    ws[uppercase_letters[col_num]+str(row_num)] = theme
    for author in output[theme].keys():
        row_num += 1
        ws[uppercase_letters[col_num]+str(row_num)] = author
        ws[uppercase_letters[col_num+1] + str(row_num)] = output[theme][author]
    col_num += 2

wb.save("TEST.xlsx")

