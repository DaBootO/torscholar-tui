import pickle
from openpyxl import Workbook
import string


def save_obj(obj, name):
    with open('obj/'+ name + '.pkl', 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)


def load_obj(name):
    with open('obj/' + name + '.pkl', 'rb') as f:
        return pickle.load(f)


uppercase_letters = list(string.ascii_uppercase)
Dict = load_obj("AuthorOutput")
output = {}


for theme in Dict.keys():
    author_dict = {}
    for query in Dict[theme].keys():
        for author in Dict[theme][query].keys():
            if author in list(author_dict.keys()):
                author_dict[author][0] += Dict[theme][query][author][0]
                author_dict[author][1] += Dict[theme][query][author][1]
            else:
                author_dict[author] = [Dict[theme][query][author][0], Dict[theme][query][author][1]]
    output[theme] = dict(sorted(author_dict.items(), key=lambda item: item[1][0], reverse=True))

wb = Workbook()
ws = wb.active

col_num = 0
for theme in output.keys():
    row_num = 1
    ws[uppercase_letters[col_num]+str(row_num)] = theme
    ws[uppercase_letters[col_num+1]+str(row_num)] = "# of articles"
    ws[uppercase_letters[col_num+2]+str(row_num)] = "# of citations"

    for author in output[theme].keys():
        row_num += 1
        ws[uppercase_letters[col_num]+str(row_num)] = author
        ws[uppercase_letters[col_num+1] + str(row_num)] = output[theme][author][0]
        ws[uppercase_letters[col_num+2] + str(row_num)] = output[theme][author][1]
    col_num += 3

wb.save("TEST.xlsx")

