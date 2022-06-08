import time
import openpyxl
import os
import string
import re
from rich import print as rprint

def dim2list(ws):
    dims = ws.dimensions.split(':')
    for i in range(len(dims)):
        dims[i] = int(re.findall('\d+', dims[i])[0])
    return dims

items = 12

letters = string.ascii_uppercase

directory = os.path.realpath(__file__)[:-len(os.path.basename(__file__))]
directory_data = os.path.join(directory, "comparison/")
# saveto = os.path.join(directory, "data/", "output/")
# data = [] # [[name, yearlow, yearhigh, positives]]

dirs = os.listdir(directory_data)
dirs = [dir for dir in dirs if os.path.isdir(os.path.join(directory_data, dir))]

files_dict = {}
for dir in dirs:
    filelist = os.listdir(os.path.join(directory_data, dir))
    files_dict[dir] = filelist

differences = 0
file_diffs = []
for dir in files_dict.keys():
    for dir_check in files_dict.keys():
        if dir == dir_check:
            continue
        for file in files_dict[dir]:
            if file not in files_dict[dir_check]:
                differences += 1
                file_diffs.append("%s is in %s but not in %s" % (file, dir, dir_check))

if differences != 0:
    print("There are some files which are not in all directories!")
    for diff in file_diffs:
        print(diff)
else:
    print("The files in all directories overlap fully!")

title_dict = {}

full_data = {}

for dir in dirs:
    title_dict[dir] = {}
    full_data[dir] = {}
    for file in os.listdir(os.path.join(directory_data, dir)):
        if file.endswith(".xlsx"):
            full_data[dir][file] = {}
            title_dict[dir][file] = []
            print("parsing titles of %s in %s" % (file, dir))
            wb = openpyxl.load_workbook(filename=os.path.join(directory_data, dir, file))
            ws = wb.active
            dims = dim2list(ws)
            index = 1
            
            while ws['C'+str(index)].value != "title":
                index += 1
            index += 1
            while index <= dims[1]:
                article = ws['C'+str(index)].value
                title_dict[dir][file].append(article)
                data = []
                for i in range(items):
                    data.append(ws[letters[i]+str(index)].value)
                full_data[dir][file][article] = data
                index += 1

already_parsed = []
for dir in title_dict.keys():
    for dir_check in title_dict.keys():
        if dir == dir_check:
            continue
        for file in title_dict[dir].keys():
            outputs = []
            title_diffs = []
            if file in title_dict[dir_check].keys():
                for title in title_dict[dir][file]:
                    if title not in title_dict[dir_check][file]:
                        if full_data[dir][file][title] not in already_parsed:
                            already_parsed.append(full_data[dir][file][title])
                        else:
                            print("We already parsed %s! skipping..." % title)
                        outputs.append(full_data[dir][file][title])
                        title_diffs.append(title)
            else:
                print("%s exists in %s but not in %s" % (file, dir, dir_check))
            if len(title_diffs) > 0:
                print("There are %i differences in titles for %s in %s and %s" % (len(title_diffs), file, dir, dir_check))
                
                out_excel = openpyxl.Workbook()
                out_ws = out_excel.active
                
                row = 1
                for i in outputs:
                    for j in range(items):
                        out_ws[letters[j]+str(row)] = i[j]
                    row += 1
                out_ws.column_dimensions['C'].width = 100.0
                out_ws.column_dimensions['D'].width = 100.0
                if not os.path.isdir(os.path.join(directory, "diffs")):
                    os.mkdir(os.path.join(directory, "diffs"))
                out_excel.save(os.path.join(os.path.join(directory, "diffs", "DIFFS"+file)))

'''
for file in os.listdir(directory_data):
    # print(file)
    positives = 0
    negatives = 0

    if file.split('.')[-1] == "xlsx":
        wb = openpyxl.load_workbook(filename=directory_data + str(file))
        ws = wb.active

        for cell in ws['A']:
            # print(cell.value)
            if cell.value == 1:
                positives += 1
            elif cell.value == 0:
                negatives += 1

        for cell in ws['B']:
            # print(cell.value)
            if cell.value == "doppelt":
                positives -= 1

        name = '_'.join(file.split('.')[0].split('_')[:-1])
        yearlow, yearhigh = file.split('.')[0].split('_')[-1].split('-')

        print("There are", positives, "in", name, "from", yearlow, "to", yearhigh, "and", negatives, "negatives")

        data.append([name, yearlow, yearhigh, positives, negatives])

wb = openpyxl.Workbook()
ws = wb.active

datapoints = len(data)

row_index = 1
col_index = 1
ws.cell(row=row_index, column=col_index).value = "article title"
col_index += 1
ws.cell(row=row_index, column=col_index).value = "yearlow"
col_index += 1
ws.cell(row=row_index, column=col_index).value = "yearhigh"
col_index += 1
ws.cell(row=row_index, column=col_index).value = "positives"
col_index += 1
ws.cell(row=row_index, column=col_index).value = "negatives"
col_index += 1
ws.cell(row=row_index, column=col_index).value = ("from - to")
row_index += 1

for dp in data:
    col_index = 1

    name = dp[0]
    yearlow = dp[1]
    yearhigh = dp[2]
    positives = dp[3]
    negatives = dp[4]

    ws.cell(row=row_index, column=col_index).value = name
    col_index += 1
    ws.cell(row=row_index, column=col_index).value = yearlow
    col_index += 1
    ws.cell(row=row_index, column=col_index).value = yearhigh
    col_index += 1
    ws.cell(row=row_index, column=col_index).value = positives
    col_index += 1
    ws.cell(row=row_index, column=col_index).value = negatives
    col_index += 1
    ws.cell(row=row_index, column=col_index).value = yearlow+'-'+yearhigh
    row_index += 1
if not os.path.exists(saveto):
    os.mkdir(saveto)
wb.save(saveto+"data_output"+ time.strftime("%d%m%y_%H%M%S") +".xlsx")

'''