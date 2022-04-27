import time
import openpyxl
import os
import string

years = [[1946, 1950],
         [1951, 1955],
         [1956, 1960],
         [1961, 1965],
         [1966, 1970],
         [1971, 1975],
         [1976, 1980],
         [1981, 1985],
         [1986, 1990],
         [1991, 1995],
         [1996, 2000],
         [2001, 2005],
         [2006, 2010],
         [2011, 2015],
         [2016, 2020],
         [2021, 2025]]

items = 12

letters = string.ascii_uppercase

directory = os.path.realpath(__file__)[:-len(os.path.basename(__file__))]
directory_data = os.path.join(directory, "data/")
saveto = os.path.join(directory, "data/", "output/")
data = [] # [[name, yearlow, yearhigh, positives]]

for file in os.listdir(directory_data):
    # print(file)
    positives = 0
    negatives = 0

    if file.split('.')[-1] == "xlsx":
        wb = openpyxl.load_workbook(filename=directory_data + str(file))
        ws = wb.active

        for cell in ws['A']:
            # print(cell.value)
            if cell.value == 1:
                positives += 1
            elif cell.value == 0:
                negatives += 1

        for cell in ws['B']:
            # print(cell.value)
            if cell.value == "doppelt":
                positives -= 1

        name = '_'.join(file.split('.')[0].split('_')[:-1])
        yearlow, yearhigh = file.split('.')[0].split('_')[-1].split('-')

        print("There are", positives, "in", name, "from", yearlow, "to", yearhigh, "and", negatives, "negatives")

        data.append([name, yearlow, yearhigh, positives, negatives])

wb = openpyxl.Workbook()
ws = wb.active

datapoints = len(data)

row_index = 1
col_index = 1
ws.cell(row=row_index, column=col_index).value = "article title"
col_index += 1
for year_list in years:
    year_string = str(year_list[0])+'-'+str(year_list[1])
    ws.cell(row=row_index, column=col_index).value = year_string
    col_index += 1
row_index += 1
#!old
# ws.cell(row=row_index, column=col_index).value = "yearlow"
# col_index += 1
# ws.cell(row=row_index, column=col_index).value = "yearhigh"
# col_index += 1
# ws.cell(row=row_index, column=col_index).value = "positives"
# col_index += 1
# ws.cell(row=row_index, column=col_index).value = "negatives"
# col_index += 1
# ws.cell(row=row_index, column=col_index).value = ("from - to")
# row_index += 1

changed_name = False
last_name = data[0][0]

for dp in data:
    col_index = 1

    name = dp[0]
    
    if last_name != name:
        row_index += 2
        last_name = name
    
    yearlow = dp[1]
    yearhigh = dp[2]
    positives = dp[3]
    negatives = dp[4]

    ws.cell(row=row_index, column=col_index).value = name + " POSITIVES"
    ws.cell(row=row_index+1, column=col_index).value = name + " NEGATIVES"
    col_index += 1
    for yl in years:
        if int(yearlow) in yl:
            col_index = years.index(yl) + 2
    ws.cell(row=row_index, column=col_index).value = positives
    ws.cell(row=row_index+1, column=col_index).value = negatives
    # ws.cell(row=row_index, column=col_index).value = name
    # col_index += 1
    # ws.cell(row=row_index, column=col_index).value = yearlow
    # col_index += 1
    # ws.cell(row=row_index, column=col_index).value = yearhigh
    # col_index += 1
    # ws.cell(row=row_index, column=col_index).value = positives
    # col_index += 1
    # ws.cell(row=row_index, column=col_index).value = negatives
    # col_index += 1
    # ws.cell(row=row_index, column=col_index).value = yearlow+'-'+yearhigh
    
if not os.path.exists(saveto):
    os.mkdir(saveto)
wb.save(saveto+"data_output"+ time.strftime("%d%m%y_%H%M%S") +".xlsx")

