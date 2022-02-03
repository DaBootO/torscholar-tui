import openpyxl
import os
import string

items = 12
# thereisfaulty = None
# newline = None
letters = string.ascii_uppercase

directory = "/home/dabooto/PycharmProjects/GoogleParser/excel/backup/"
saveto = "/home/dabooto/PycharmProjects/GoogleParser/excel/backup/finished/"
for file in os.listdir(directory):
    # print(file)
    if file.split('.')[-1] == "xlsx":
        print(file, "is being processed...")
        wb = openpyxl.load_workbook(filename=directory+str(file))
        ws = wb.active
        rows = len(tuple(ws.rows))
        columns = len(tuple(ws.columns))
        pos = str('A') + str(1) + ":" + str(letters[columns+1]) + str(rows)
        ws.move_range(pos, rows=0, cols=2)
        ws['A1'] = "0/1"
        ws['B1'] = "comment"
        ws.column_dimensions['C'].width = 100.0
        ws.column_dimensions['D'].width = 100.0
        wb.save(saveto+str(file))
