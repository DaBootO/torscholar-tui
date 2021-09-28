import os
import string
import openpyxl

items = 12
directory = "/home/dabooto/PycharmProjects/GoogleParser/january/data/"

OLD_list = os.listdir(directory+"OLD")
NEW_list = os.listdir(directory+"NEW")

letters = string.ascii_uppercase

for file in OLD_list:
    if file in NEW_list:
        print(file, "exists in OLD and NEW!")
        wb = openpyxl.load_workbook(filename=directory + "OLD/" + str(file))
        ws = wb.active
        row = 1
        while ws['C'+str(row)].value != "title":
            row += 1
        row += 1
        print("data start in row:", row)

        OLD_paper_list = []

        while ws["C"+str(row)].value is not None:
            paper = ws["C"+str(row)].value
            OLD_paper_list.append(paper)
            row += 1

        # print(OLD_paper_list)

        wb.close()

        wb = openpyxl.load_workbook(filename=directory + "NEW/" + str(file))
        ws = wb.active
        row = 1
        while ws['C' + str(row)].value != "title":
            row += 1
        row += 1
        print("data start in row:", row)

        NEW_paper_list = []

        while ws["C" + str(row)].value is not None:
            paper = ws["C" + str(row)].value
            NEW_paper_list.append(paper)
            row += 1

        # print(NEW_paper_list)

        wb.close()

        print("length of OLD_paper_list:", len(OLD_paper_list))
        print("length of NEW_paper_list:", len(NEW_paper_list))

        VERYNEW_papers = []

        for paper in NEW_paper_list:
            if paper not in OLD_paper_list:
                VERYNEW_papers.append(paper)
        print("REST:", len(VERYNEW_papers))
        # print(VERYNEW_papers)

        wb = openpyxl.load_workbook(filename=directory + "NEW/" + str(file))
        ws = wb.active
        row = 1
        while ws['C' + str(row)].value != "title":
            row += 1
        row += 1

        while ws['C' + str(row)].value is not None:
            if ws['C' + str(row)].value not in VERYNEW_papers:
                # print("DELETED!")
                ws.delete_rows(row)
            else:
                row += 1

        wb.save(directory + "NEW/" + str(file))






    # print(file)
    # if file.split('.')[-1] == "xlsx":
    #     print(file, "is being processed...")
    #     wb = openpyxl.load_workbook(filename=directory+str(file))
    #     ws = wb.active
    #     rows = len(tuple(ws.rows))
    #     columns = len(tuple(ws.columns))
    #     pos = str('A') + str(1) + ":" + str(letters[columns+1]) + str(rows)
    #     ws.move_range(pos, rows=0, cols=2)
    #     ws['A1'] = "0/1"
    #     ws['B1'] = "comment"
    #     ws.column_dimensions['C'].width = 100.0
    #     ws.column_dimensions['D'].width = 100.0
    #     wb.save(saveto+str(file))