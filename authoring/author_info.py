import os
import openpyxl
import re
import sys
from tqdm import trange

def dim2list(ws):
    dims = ws.dimensions.split(':')
    for i in range(len(dims)):
        dims[i] = int(re.findall('\d+', dims[i])[0])
    return dims

base_directory = "/home/dabooto/PycharmProjects/AuthorData/wo_doubles"

files = os.listdir(base_directory)
print(files)

file_number = 0

author_search = sys.argv[1]

for f in files:

    if f[-4:] == 'xlsx':
        wb = openpyxl.load_workbook(base_directory+'/'+f)
        ws = wb.active
        dims = dim2list(ws)

        row = 1

        for row_a in range(row, dims[1]+1):
            line = ws['D'+str(row_a)].value
            article = ws['C'+str(row_a)].value
            is_part_of = ws['A'+str(row_a)].value
            comment = ws['B'+str(row_a)].value
            citations = ws['G'+str(row_a)].value
            if type(citations) is str and citations != "None":
                try:
                    citations = int(citations)
                except Exception:
                    print(f, row_a)
                    print("something went wrong in the aforementioned sheet")
                    print(type(citations))
            else:
                citations = 0

            if line is not None:
                if is_part_of == 1 and comment != "doppelt":
                    if '-' in line:
                        auth = line[:line.index('-')]
                    else:
                        auth = line
                else:
                    continue
            else:
                continue

            auth = auth.strip().upper()
            auth = re.sub(r'[^A-Z\s,]+', '', auth)
            auth = auth.split(',')
            for authi in auth:
                authi = authi.strip()
                if author_search.lower() in authi.lower():
                    print(article[:60])
                    print(line[:60])
                    print(citations)
        wb.close()
