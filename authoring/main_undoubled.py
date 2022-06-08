import os
import openpyxl
import re
import pickle
from tqdm import trange


def save_obj(obj, directory ,name):
    if not os.path.exists(os.path.join(directory, "obj/")):
            os.mkdir(os.path.join(directory, "obj/"))
    with open(os.path.join(directory, "obj/"+ name + '.pkl'), 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)


def load_obj(name, directory):
    with open(os.path.join(directory, "obj/"+ name + '.pkl'), 'rb') as f:
        return pickle.load(f)


def dim2list(ws):
    dims = ws.dimensions.split(':')
    for i in range(len(dims)):
        dims[i] = int(re.findall('\d+', dims[i])[0])
    return dims

# base_directory = "/home/dabooto/PycharmProjects/AuthorData/wo_doubles"
directory = os.path.abspath(os.path.realpath(__file__)[:-len(os.path.basename(__file__))]) + '/'
base_directory = os.path.join(directory, "wo_doubles")

files = os.listdir(base_directory)

file_number = 0
DataOut = {}

for f in files:
    authors = {}

    if f[-4:] == 'xlsx':
        DataOut[f[:-5]] = {}
        print("parsing %s..." % f)
        wb = openpyxl.load_workbook(base_directory+'/'+f)
        ws = wb.active
        dims = dim2list(ws)

        row = 1

        for row_a in trange(row, dims[1]+1):
            line = ws['D'+str(row_a)].value
            is_part_of = ws['A'+str(row_a)].value
            comment = ws['B'+str(row_a)].value
            citations = ws['G'+str(row_a)].value
            if type(citations) is str and citations != "None":
                try:
                    citations = int(citations)
                except Exception:
                    print("something went wrong in the aforementioned sheet")
                    print(f, row_a)
                    print(type(citations))
            else:
                citations = 0

            if line is not None:
                if is_part_of == 1 and comment != "doppelt":
                    if '-' in line:
                        auth = line[:line.index('-')]
                    else:
                        auth = line
                else:
                    continue
            else:
                continue

            #
            # if line != None and '-' in line:
            #     auth = line[:line.index('-')]
            # else:
            #     if line != None and ',' in line:
            #         auth = line
            #     else:
            #         continue

            auth = auth.strip().upper()
            auth = re.sub(r'[^A-Z\s,]+', '', auth)
            auth = auth.split(',')
            for authi in auth:
                authi = authi.strip()
                if authi != "":
                    if authi in authors:
                        authors[authi][0] += 1
                        authors[authi][1] += citations
                    else:
                        authors[authi] = [1, citations]
        filename = f[:-5]
        DataOut[f[:-5]] = authors
        wb.close()


save_obj(DataOut, directory, "AuthorOutputUndoubled")