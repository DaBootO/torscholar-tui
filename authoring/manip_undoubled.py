import pickle
from openpyxl import Workbook
import string
import os
from tqdm import tqdm

def save_obj(obj, directory ,name):
    with open(os.path.join(directory, "obj/"+ name + '.pkl'), 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)


def load_obj(name, directory):
    with open(os.path.join(directory, "obj/"+ name + '.pkl'), 'rb') as f:
        return pickle.load(f)

directory = os.path.abspath(os.path.realpath(__file__)[:-len(os.path.basename(__file__))]) + '/'

uppercase_letters = list(string.ascii_uppercase)
Dict = load_obj("AuthorOutputUndoubled", directory)
output = {}


for query in Dict.keys():
    author_dict = {}
    for author in tqdm(Dict[query].keys(), bar_format='{percentage:3.0f}%'):
        if author in list(author_dict.keys()):
            author_dict[author][0] += Dict[query][author][0]
            author_dict[author][1] += Dict[query][author][1]
        else:
            author_dict[author] = [Dict[query][author][0], Dict[query][author][1]]
    output[query] = dict(sorted(author_dict.items(), key=lambda item: item[1][0], reverse=True))

print("""
###########################
# finished author parsing #
###########################

waiting for excel file to be written...""")

wb = Workbook()
ws = wb.active

col_num = 0
for query in output.keys():
    row_num = 1
    ws[uppercase_letters[col_num]+str(row_num)] = query
    ws[uppercase_letters[col_num+1]+str(row_num)] = "# of articles"
    ws[uppercase_letters[col_num+2]+str(row_num)] = "# of citations"

    for author in output[query].keys():
        row_num += 1
        ws[uppercase_letters[col_num]+str(row_num)] = author
        ws[uppercase_letters[col_num+1] + str(row_num)] = output[query][author][0]
        ws[uppercase_letters[col_num+2] + str(row_num)] = output[query][author][1]
    col_num += 3

wb.save(os.path.join(directory, "Undoubled_Authors_with_citations.xlsx"))
print("Excel file written!")

