import os
import openpyxl
import re
import string
from tqdm import tqdm

def dim2list(ws):
    dims = ws.dimensions.split(':')
    for i in range(len(dims)):
        dims[i] = int(re.findall('\d+', dims[i])[0])
    return dims

directory = directory = os.path.abspath(os.path.realpath(__file__)[:-len(os.path.basename(__file__))]) + '/'
base_directory = os.path.join(directory, "Rohdaten") + '/'
undoubles_directory = os.path.join(directory, "wo_doubles") + '/'

sub_dirs = os.listdir(base_directory)
print(sub_dirs)

uppercase_letters = list(string.ascii_uppercase)

file_number = 0
DataOut = {}

wb_doubles_out = openpyxl.Workbook()
ws_doubles_out = wb_doubles_out.active
row_doubles_out = 1
ws_doubles_out['A' + str(row_doubles_out)].value = "article name"
ws_doubles_out['B' + str(row_doubles_out)].value = "authors"
ws_doubles_out['C' + str(row_doubles_out)].value = "taken from * file"
row_doubles_out += 1

for sd in sub_dirs:
    if "." not in sd:
        undoubled_list = []

        wb_ud = openpyxl.Workbook()
        ws_ud = wb_ud.active
        row_ud = 1

        sub_dir_path = base_directory + '/' + sd
        files = os.listdir(sub_dir_path)

        num_file = 1

        for f in tqdm(files, bar_format='{percentage:3.0f}%'):
            if f[-4:] == 'xlsx':
                wb = openpyxl.load_workbook(sub_dir_path+'/'+f)
                ws = wb.active
                dims = dim2list(ws)

            row = 1
            while ws['D' + str(row)].value != 'author':
                row += 1

            for row_a in range(row + 1, dims[1] + 1):

                if ws['C' + str(row_a)].value in undoubled_list:
                    ws_doubles_out['A' + str(row_doubles_out)].value = ws['C' + str(row_a)].value
                    ws_doubles_out['B' + str(row_doubles_out)].value = ws['D' + str(row_a)].value
                    ws_doubles_out['C' + str(row_doubles_out)].value = f
                    row_doubles_out += 1
                elif ws['B' + str(row_a)].value == "doppelt":
                    continue
                else:
                    undoubled_list.append(ws['C' + str(row_a)].value)
                    for column_copy in range(15):
                        ws_ud[uppercase_letters[column_copy] + str(row_ud)].value = ws[uppercase_letters[column_copy] + str(row_a)].value
                    row_ud += 1

        wb_ud.save(undoubles_directory+str(sd)+".xlsx")
        wb_doubles_out.save(undoubles_directory+"doubled_output.xlsx")







