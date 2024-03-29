import time
import openpyxl
import os
import string
import re
from rich import print as rprint
import shutil
import logging
import datetime

def dim2list(ws):
    dims = ws.dimensions.split(':')
    for i in range(len(dims)):
        dims[i] = int(re.findall('\d+', dims[i])[0])
    return dims

items = 12

letters = string.ascii_uppercase

directory = os.path.abspath(os.path.realpath(__file__)[:-len(os.path.basename(__file__))] + "../../") + '/'
directory_data = os.path.join(directory, "comparison/")
# saveto = os.path.join(directory, "data/", "output/")
# data = [] # [[name, yearlow, yearhigh, positives]]

# configuring the logger
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)-4s %(processName)s %(message)s", 
    datefmt="%H:%M:%S",
    filename=os.path.join(directory, 'comparator.log'),
)

dirs = os.listdir(directory_data)
dirs = [dir for dir in dirs if os.path.isdir(os.path.join(directory_data, dir))]

print("exploring all the files...")
logging.info("exploring all the files...")

files_dict = {}
for dir in dirs:
    filelist = os.listdir(os.path.join(directory_data, dir))
    files_dict[dir] = filelist

differences = 0
file_diffs = []
for dir in files_dict.keys():
    for dir_check in files_dict.keys():
        if dir == dir_check:
            continue
        for file in files_dict[dir]:
            if file not in files_dict[dir_check]:
                differences += 1
                file_diffs.append("%s is in %s but not in %s" % (file, dir, dir_check))

if differences != 0:
    print("There are some files which are not in all directories!")
    logging.info("There are some files which are not in all directories!")
    for diff in file_diffs:
        logging.info(diff)
else:
    print("The files in all directories overlap fully!")

title_dict = {}

print("parsing the whole dataset...")
logging.info("parsing the whole dataset...")
# just making a dict with every file and its contents
full_data = {}
for dir in dirs:
    title_dict[dir] = {}
    full_data[dir] = {}
    for file in os.listdir(os.path.join(directory_data, dir)):
        if file.endswith(".xlsx"):
            full_data[dir][file] = {}
            title_dict[dir][file] = []
            logging.info("parsing titles of %s in %s" % (file, dir))
            wb = openpyxl.load_workbook(filename=os.path.join(directory_data, dir, file))
            ws = wb.active
            dims = dim2list(ws)
            index = 1
            
            while ws['C'+str(index)].value != "title":
                index += 1
            index += 1
            while index <= dims[1]:
                article = ws['C'+str(index)].value
                title_dict[dir][file].append(article)
                data = []
                for i in range(items):
                    data.append(ws[letters[i]+str(index)].value)
                full_data[dir][file][article] = data
                index += 1

total_diffs = 0
print("starting checks...")
logging.info("starting checks...")
# checking the files and deciding where to put them
already_parsed = []
already_parsed_in_both = []
for dir in title_dict.keys():
    for dir_check in title_dict.keys():
        if dir == dir_check:
            continue
        for file in title_dict[dir].keys():
            outputs = []
            title_diffs = []
            similar_outputs = []
            similar_outputs_not_coded = []
            unique_uncoded = []
            doubled_inputs = 0
            if file in title_dict[dir_check].keys(): # if not -> send to unique
                for title in title_dict[dir][file]:
                    if title not in title_dict[dir_check][file]: # if title is only in one of the 2 files
                        if title not in already_parsed:
                            already_parsed.append(title)
                        else: # automatically denies entry for doubled inputs
                            doubled_inputs += 1
                            continue
                        outputs.append(full_data[dir][file][title])
                        
                        # if the title only exists in one file and is not coded -> NOT_CODED
                        if full_data[dir][file][title][0] == None:
                            unique_uncoded.append(full_data[dir][file][title])
                            logging.info("The title %s exists only in %s and is NOT coded!" % (title, file))
                        else:
                            title_diffs.append(full_data[dir][file][title])
                    else: # title is in both files
                        if title not in already_parsed_in_both:
                            already_parsed_in_both.append(title)
                        else:
                            continue
                        # check if one is coded
                        if full_data[dir][file][title][0] != None or full_data[dir_check][file][title][0] != None:
                            # if at least one is coded -> check which one has the latest modification date
                            dt_dir = datetime.datetime.fromtimestamp(os.path.getmtime(os.path.join(directory_data, dir, file)))
                            dt_dircheck = datetime.datetime.fromtimestamp(os.path.getmtime(os.path.join(directory_data, dir_check, file)))
                            
                            # quick checks to find out which title is coded (if both - check if equal)
                            if full_data[dir][file][title][0] == full_data[dir_check][file][title][0]:
                                # if equal -> newest for updating citations etc.
                                if dt_dir > dt_dircheck:
                                    title_code = full_data[dir][file][title][0]
                                    full_data[dir][file][title][0] = title_code
                                    similar_title = full_data[dir][file][title]
                                else:
                                    title_code = full_data[dir_check][file][title][0]
                                    full_data[dir_check][file][title][0] = title_code
                                    similar_title = full_data[dir_check][file][title]
                                similar_outputs.append(similar_title)
                            else:
                                # not equal
                                # quick log
                                logging.info("There is a difference of code for %s in %s" % (title, file))
                                logging.info("code diff for %s in %s" % (title, file))
                                logging.info("%s has code: %s" % (dir, full_data[dir][file][title][0]))
                                logging.info("%s has code: %s" % (dir_check, full_data[dir_check][file][title][0]))
                                
                                # if one is coded -> keep citation data from newest and set code to the existing one
                                if full_data[dir][file][title][0] != None and full_data[dir_check][file][title][0] == None:
                                    if dt_dir > dt_dircheck:
                                        full_data[dir][file][title][0] = full_data[dir][file][title][0]
                                        similar_title = full_data[dir][file][title]
                                    else:
                                        full_data[dir_check][file][title][0] = full_data[dir][file][title][0]
                                        similar_title = full_data[dir_check][file][title]
                                    similar_outputs.append(similar_title)
                                elif full_data[dir][file][title][0] == None and full_data[dir_check][file][title][0] != None:
                                    if dt_dir > dt_dircheck:
                                        full_data[dir][file][title][0] = full_data[dir_check][file][title][0]
                                        similar_title = full_data[dir][file][title]
                                    else:
                                        full_data[dir_check][file][title][0] = full_data[dir_check][file][title][0]
                                        similar_title = full_data[dir_check][file][title]
                                    similar_outputs.append(similar_title)
                                else: # only the case if code is different in both files -> better to code it again!
                                    if dt_dir > dt_dircheck:
                                        full_data[dir][file][title][0] = None
                                        similar_title_not_coded = full_data[dir][file][title]
                                    else:
                                        full_data[dir_check][file][title][0] = None
                                        similar_title_not_coded = full_data[dir_check][file][title]
                                    title_diffs.append(similar_title_not_coded)
                                
                        else: # if not coded put into special file with SUFFIX NOT_CODED
                            similar_title_not_coded = full_data[dir][file][title]
                            similar_outputs_not_coded.append(similar_title_not_coded)
                logging.info("saved time by throwing out %i doubled inputs!" % doubled_inputs)
            
            # if the file is only in one directory -> put in unique
            else:
                logging.info("%s exists in %s but not in %s" % (file, dir, dir_check))
                if not os.path.isdir(os.path.join(directory, "unique")):
                    os.mkdir(os.path.join(directory, "unique"))
                
                
                file_src = os.path.join(directory_data, dir, file)
                file_out = os.path.join(directory, "unique", file)
                shutil.copy(file_src, file_out)
            if len(title_diffs) > 0: # put the differences into a special directory
                total_diffs += len(title_diffs)
                logging.info("There are %i differences in titles for %s in %s and %s" % (len(title_diffs), file, dir, dir_check))
                
                out_excel = openpyxl.Workbook()
                out_ws = out_excel.active
                
                row = 1
                labels = [
                "0/1",
                "comment",
                "title",
                "author",
                "url",
                "year",
                "num_citations",
                "num_versions",
                "cluster_id",
                "url_pdf",
                "url_citations",
                "url_versions",
                "url_citation",
                "excerpt"]

                for lbl in range(len(labels)):
                    out_ws[letters[lbl]+str(row)] = labels[lbl]
                row += 1
                for i in title_diffs:
                    for j in range(items):
                        out_ws[letters[j]+str(row)] = i[j]
                    row += 1
                out_ws.column_dimensions['C'].width = 100.0
                out_ws.column_dimensions['D'].width = 100.0
                if not os.path.isdir(os.path.join(directory, "diffs")):
                    os.mkdir(os.path.join(directory, "diffs"))
                out_excel.save(os.path.join(os.path.join(directory, "diffs", "DIFFS"+file)))
            # work with the similarities
            if len(similar_outputs) > 0:
                out_excel = openpyxl.Workbook()
                out_ws = out_excel.active
                
                row = 1
                labels = [
                "0/1",
                "comment",
                "title",
                "author",
                "url",
                "year",
                "num_citations",
                "num_versions",
                "cluster_id",
                "url_pdf",
                "url_citations",
                "url_versions",
                "url_citation",
                "excerpt"]
                
                for lbl in range(len(labels)):
                    out_ws[letters[lbl]+str(row)] = labels[lbl]
                row += 1
                for i in similar_outputs:
                    for j in range(items):
                        out_ws[letters[j]+str(row)] = i[j]
                    row += 1
                out_ws.column_dimensions['C'].width = 100.0
                out_ws.column_dimensions['D'].width = 100.0
                if not os.path.isdir(os.path.join(directory, "similar")):
                    os.mkdir(os.path.join(directory, "similar"))
                out_excel.save(os.path.join(os.path.join(directory, "similar", file)))
            if len(similar_outputs_not_coded) > 0:
                out_excel = openpyxl.Workbook()
                out_ws = out_excel.active
                
                row = 1
                labels = [
                "0/1",
                "comment",
                "title",
                "author",
                "url",
                "year",
                "num_citations",
                "num_versions",
                "cluster_id",
                "url_pdf",
                "url_citations",
                "url_versions",
                "url_citation",
                "excerpt"]
                
                for lbl in range(len(labels)):
                    out_ws[letters[lbl]+str(row)] = labels[lbl]
                row += 1
                for i in similar_outputs_not_coded:
                    for j in range(items):
                        out_ws[letters[j]+str(row)] = i[j]
                    row += 1
                out_ws.column_dimensions['C'].width = 100.0
                out_ws.column_dimensions['D'].width = 100.0
                if not os.path.isdir(os.path.join(directory, "similar")):
                    os.mkdir(os.path.join(directory, "similar"))
                out_excel.save(os.path.join(os.path.join(directory, "similar", "NOT_CODED"+file)))
            if len(unique_uncoded) > 0:
                out_excel = openpyxl.Workbook()
                out_ws = out_excel.active
                
                row = 1
                labels = [
                "0/1",
                "comment",
                "title",
                "author",
                "url",
                "year",
                "num_citations",
                "num_versions",
                "cluster_id",
                "url_pdf",
                "url_citations",
                "url_versions",
                "url_citation",
                "excerpt"]
                
                for lbl in range(len(labels)):
                    out_ws[letters[lbl]+str(row)] = labels[lbl]
                row += 1
                for i in unique_uncoded:
                    for j in range(items):
                        out_ws[letters[j]+str(row)] = i[j]
                    row += 1
                out_ws.column_dimensions['C'].width = 100.0
                out_ws.column_dimensions['D'].width = 100.0
                if not os.path.isdir(os.path.join(directory, "diffs")):
                    os.mkdir(os.path.join(directory, "diffs"))
                out_excel.save(os.path.join(os.path.join(directory, "diffs", "NOT_CODED"+file)))
                # print("There are no differences in titles for %s in %s and %s" % (file, dir, dir_check))
                # if not os.path.isdir(os.path.join(directory, "similar")):
                #     os.mkdir(os.path.join(directory, "similar"))
                # file_src = os.path.join(directory_data, dir, file)
                # file_out = os.path.join(directory, "similar", file)
                # shutil.copy(file_src, file_out)

print("There is a total of %i differences!" % total_diffs)
logging.info("There is a total of %i differences!" % total_diffs)