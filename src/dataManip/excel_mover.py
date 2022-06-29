import openpyxl
import os
import string
import sys

items = 12
# thereisfaulty = None
# newline = None
letters = string.ascii_uppercase

directory = os.path.abspath(os.path.realpath(__file__)[:-len(os.path.basename(__file__))] + "../../") + '/'
directory_excel = os.path.join(directory, "excel/")
if not os.path.isdir(directory_excel):
    os.mkdir(directory_excel)

saveto = os.path.join(directory, "excel/", "finished/")
if not os.path.isdir(saveto):
    os.mkdir(saveto)

for file in os.listdir(directory_excel):
    # print(file)
    if file.split('.')[-1] == "xlsx":
        print(file, "is being processed...")
        wb = openpyxl.load_workbook(filename=directory_excel+str(file))
        ws = wb.active
        rows = len(tuple(ws.rows))
        columns = len(tuple(ws.columns))
        pos = str('A') + str(1) + ":" + str(letters[columns+1]) + str(rows)
        ws.move_range(pos, rows=0, cols=2)
        ws['A1'] = "0/1"
        ws['B1'] = "comment"
        ws.column_dimensions['C'].width = 100.0
        ws.column_dimensions['D'].width = 100.0
        if not os.path.exists(saveto):
                os.mkdir(saveto)
        wb.save(saveto+str(file))
