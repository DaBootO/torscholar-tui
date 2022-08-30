import time
import openpyxl
import os
import string
import re
from rich import print as rprint
import shutil
import logging

def dim2list(ws):
    dims = ws.dimensions.split(':')
    for i in range(len(dims)):
        dims[i] = int(re.findall('\d+', dims[i])[0])
    return dims


# how many columns there are in the excel files
items = 12

# needed for easier access to cells in excel
letters = string.ascii_uppercase

# just creating the path strings
directory = os.path.abspath(os.path.realpath(__file__)[:-len(os.path.basename(__file__))] + "../../") + '/'
similar_data_path = os.path.join(directory, "similar/")
unique_data_path = os.path.join(directory, "unique/")
diff_data_path = os.path.join(directory, "diffs/")

if not os.path.isdir(os.path.join(directory, "merged")):
    os.mkdir(os.path.join(directory, "merged"))

# configuring the logger
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)-4s %(processName)s %(message)s", 
    datefmt="%H:%M:%S",
    filename=os.path.join(directory, 'merger.log'),
)

# copying all the unique files to the merged directory
# no checks needed, as unique files do not need to be compared
if os.path.isdir(unique_data_path):
    for unique_file in os.listdir(unique_data_path):
        file_src = os.path.join(unique_data_path, unique_file)
        file_out = os.path.join(directory, "merged", unique_file)
        logging.info("unique file %s is being transfered..." % unique_file)
        shutil.copy(file_src, file_out)

# listing files in the similar and diff directories
if os.path.isdir(similar_data_path):
    similar_data = os.listdir(similar_data_path)
    similar_files_dict = {}
if os.path.isdir(diff_data_path):
    diff_data = os.listdir(diff_data_path)
    diff_files_dict = {}


merge_files = []

if os.path.isdir(diff_data_path):
    for diff_file in diff_data:
        if diff_file.replace('DIFFS', '') in similar_data: # if DIFFS_XXX and XXX exist -> append to merge list
            merge_files.append(diff_file)
        else:
            shutil.move(os.path.join(diff_data_path, diff_file),
                        os.path.join(similar_data_path, diff_file.replace('DIFFS', 'NOT_CODED')))
if os.path.isdir(similar_data_path):
    for similar_file in similar_data:
        if 'DIFFS'+similar_file not in merge_files: # if similar file exists without its DIFF counterpart -> copy over
            file_src = os.path.join(similar_data_path, similar_file)
            file_out = os.path.join(directory, "merged", similar_file)
            logging.info("similar file %s is being transfered..." % similar_file)
            if not os.path.isdir(os.path.join(directory, "merged/")):
                os.mkdir(os.path.join(directory, "merged/"))
            shutil.copy(file_src, file_out)

print("There are %s files to be merged left!" % len(merge_files))

for merger in merge_files:
    diff_file_path = os.path.join(diff_data_path, merger)
    similar_file_path = os.path.join(similar_data_path, merger.replace('DIFFS',''))
    #! debug
    # print(diff_file_path)
    # print(similar_file_path)
    wb_sim = openpyxl.load_workbook(filename=similar_file_path)
    ws_sim = wb_sim.active
    # find end of similar file
    row_sim = 1
    while not isinstance(ws_sim['C'+str(row_sim)].value, type(None)):
        row_sim += 1
    #! debug
    # print(ws_sim['C'+str(row_sim)].value)
    # print(ws_sim['C'+str(row_sim-1)].value)
    wb_diff = openpyxl.load_workbook(filename=diff_file_path)
    ws_diff = wb_diff.active
    # find start of similar file
    row_diff = 1
    while ws_diff['C'+str(row_diff)].value != "title":
        row_diff += 1
    row_diff += 1
    #! debug
    # print(ws_diff['C'+str(row_diff)].value)
    # print(ws_diff['C'+str(row_diff+1)].value)
    for row_diff_copy in range(dim2list(ws_diff)[1]-1):
        for i in range(items):
            ws_sim[letters[i]+str(row_sim)].value = ws_diff[letters[i]+str(row_diff + row_diff_copy)].value
        row_sim += 1
    logging.info("merged file %s is being transfered..." % merger.replace('DIFFS',''))
    wb_sim_out = file_out = os.path.join(directory, "merged", merger.replace('DIFFS',''))
    wb_sim.save(wb_sim_out)
'''
differences = 0
file_diffs = []
for dir in files_dict.keys():
    for dir_check in files_dict.keys():
        if dir == dir_check:
            continue
        for file in files_dict[dir]:
            if file not in files_dict[dir_check]:
                differences += 1
                file_diffs.append("%s is in %s but not in %s" % (file, dir, dir_check))

if differences != 0:
    print("There are some files which are not in all directories!")
    for diff in file_diffs:
        print(diff)
else:
    print("The files in all directories overlap fully!")

title_dict = {}

full_data = {}

for dir in dirs:
    title_dict[dir] = {}
    full_data[dir] = {}
    for file in os.listdir(os.path.join(directory_data, dir)):
        if file.endswith(".xlsx"):
            full_data[dir][file] = {}
            title_dict[dir][file] = []
            print("parsing titles of %s in %s" % (file, dir))
            wb = openpyxl.load_workbook(filename=os.path.join(directory_data, dir, file))
            ws = wb.active
            dims = dim2list(ws)
            index = 1
            
            while ws['C'+str(index)].value != "title":
                index += 1
            index += 1
            while index <= dims[1]:
                article = ws['C'+str(index)].value
                title_dict[dir][file].append(article)
                data = []
                for i in range(items):
                    data.append(ws[letters[i]+str(index)].value)
                full_data[dir][file][article] = data
                index += 1

already_parsed = []
for dir in title_dict.keys():
    for dir_check in title_dict.keys():
        if dir == dir_check:
            continue
        for file in title_dict[dir].keys():
            outputs = []
            title_diffs = []
            if file in title_dict[dir_check].keys():
                for title in title_dict[dir][file]:
                    if title not in title_dict[dir_check][file]:
                        if full_data[dir][file][title] not in already_parsed:
                            already_parsed.append(full_data[dir][file][title])
                        else:
                            print("We already parsed %s! skipping..." % title)
                        outputs.append(full_data[dir][file][title])
                        title_diffs.append(title)
            else:
                print("%s exists in %s but not in %s" % (file, dir, dir_check))
                if not os.path.isdir(os.path.join(directory, "unique")):
                    os.mkdir(os.path.join(directory, "unique"))
                file_src = os.path.join(directory_data, dir, file)
                file_out = os.path.join(directory, "unique", file)
                shutil.copy(file_src, file_out)
                continue
            if len(title_diffs) > 0:
                print("There are %i differences in titles for %s in %s and %s" % (len(title_diffs), file, dir, dir_check))
                
                out_excel = openpyxl.Workbook()
                out_ws = out_excel.active
                
                row = 1
                labels = [
                "0/1",
                "comment",
                "title",
                "author",
                "url",
                "year",
                "num_citations",
                "num_versions",
                "cluster_id",
                "url_pdf",
                "url_citations",
                "url_versions",
                "url_citation",
                "excerpt"]

                for lbl in range(len(labels)):
                    out_ws[letters[lbl]+str(row)] = labels[lbl]
                row += 1
                for i in outputs:
                    for j in range(items):
                        out_ws[letters[j]+str(row)] = i[j]
                    row += 1
                out_ws.column_dimensions['C'].width = 100.0
                out_ws.column_dimensions['D'].width = 100.0
                if not os.path.isdir(os.path.join(directory, "diffs")):
                    os.mkdir(os.path.join(directory, "diffs"))
                out_excel.save(os.path.join(os.path.join(directory, "diffs", "DIFFS"+file)))
            else:
                print("There are no differences in titles for %s in %s and %s" % (file, dir, dir_check))
                if not os.path.isdir(os.path.join(directory, "similar")):
                    os.mkdir(os.path.join(directory, "similar"))
                file_src = os.path.join(directory_data, dir, file)
                file_out = os.path.join(directory, "similar", file)
                shutil.copy(file_src, file_out)
'''

print("merging complete!")